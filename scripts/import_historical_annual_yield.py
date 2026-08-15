#!/usr/bin/env python3
"""One-off import of even-older historical harvest data (1987-2019) into the
HistoricalAnnualYield table, extending the farm's history further back than
HistoricalHarvest's daily records (which only go back to 2020).

Two sources, both in "Actual farm data/OES 2025 21 Nov.xlsx":

- "Produksie" sheet, 2012-2019: the farm's own annual PER-BLOCK production
  summary, using today's block ids. Unlike the daily 2020-2025 workbook,
  this only has one total per block per season, not per day. Years
  2012-2015 are block 7 only (the other blocks weren't yet in production -
  the sheet's own "Bome"/tree-count and "Plant dat"/plant-year rows confirm
  this).

- "kg per blok" sheet, 1987-2009: even older records that predate today's
  block register entirely - the sheet's own block columns use a totally
  different, incompatible numbering (Blok 1-14, 2b/3b/4b/7b, Lw1a/Lw1b)
  with no reliable mapping to today's blocks. Rather than guess at that
  mapping, only the sheet's own farm-wide "Totaal" (kg picked) column is
  imported here, as a single block_id=None row per year - a whole-farm
  total with no block breakdown. 2010-2011 have no data in this sheet at
  all (blank rows - not imported), and 2012+ is better covered by
  "Produksie" above.

2020 onward is skipped entirely here since HistoricalHarvest already covers
those years at daily granularity from a better source.

Block splits (Produksie years only): like the daily import, the sheet's
"8a", "10", "17" and "19" columns are actually pre-split combined totals
for old blocks 8/10/17/19 (confirmed by cross-checking against
HistoricalHarvest's 2020-2022 daily sums, which match these columns almost
exactly). Split between today's sub-blocks by the same hectare ratios used
for the daily import - see scripts/import_historical_harvest.py's
docstring for the ratios.

Safe to re-run: replaces the whole table each time.

Usage (run with the backend's own venv so sqlmodel etc. are on the path):
    backend/.venv/bin/python3 scripts/import_historical_annual_yield.py
"""
import os
import sys

import openpyxl
from openpyxl.utils import column_index_from_string

BACKEND_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "backend")
sys.path.insert(0, BACKEND_DIR)

from sqlmodel import Session, delete  # noqa: E402

from db import create_db_and_tables, engine  # noqa: E402
from models import HistoricalAnnualYield  # noqa: E402

XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Actual farm data",
                          "OES 2025 21 Nov.xlsx")
PRODUKSIE_SHEET = "Produksie"
PRODUKSIE_YEARS = range(2012, 2020)  # 2020+ already covered by HistoricalHarvest at daily granularity

# column letter (on the "Produksie" sheet) -> block id as recorded there.
# "8a", "10", "17", "19" are pre-split combined totals for old blocks 8/10/17/19.
COLUMN_BLOCKS = {
    "C": "7", "E": "8a", "G": "9", "I": "10", "K": "11", "M": "12", "O": "13",
    "Q": "14", "S": "15", "U": "16", "W": "17", "Y": "18", "AA": "19",
}
SPLIT_RATIOS = {  # old combined block id -> {new sub-block id: share of hectares}
    "8a": {"8a": 0.583, "8b": 0.417},
    "10": {"10a": 0.510, "10b": 0.490},
    "17": {"17a": 0.500, "17b": 0.500},
    "19": {"19a": 0.500, "19b": 0.500},
}

KG_PER_BLOK_SHEET = "kg per blok"
KG_PER_BLOK_YEARS = range(1987, 2010)  # 2010-2011 blank in the source; 2012+ covered by Produksie above
KG_PER_BLOK_YEAR_COL = "B"
KG_PER_BLOK_TOTAAL_COL = "W"  # whole-farm kg picked - old block columns aren't mapped to today's blocks
# The clean "Jaar" table this data comes from is rows 6-28. Below that (from row ~35 on) the sheet
# has a second, informal "Skatting" (estimate) scratch area re-deriving some of the same years by
# hand - row/column-misaligned, labelled with things like "Skat kg/boom" and a name ("Charl"), and
# disagreeing with the clean table's own numbers for the years they overlap. Not scanning past the
# clean table avoids picking up those unreliable duplicates.
KG_PER_BLOK_MIN_ROW, KG_PER_BLOK_MAX_ROW = 6, 28


def load_produksie_rows(wb):
    ws = wb[PRODUKSIE_SHEET]
    year_col_idx = column_index_from_string("B") - 1  # 0-based index into each row tuple
    value_cols = {column_index_from_string(letter) - 1: block_id for letter, block_id in COLUMN_BLOCKS.items()}
    rows = []
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
        year = row[year_col_idx]
        if not isinstance(year, int) or year not in PRODUKSIE_YEARS:
            continue
        for col_idx, block_id in value_cols.items():
            kg = row[col_idx]
            if kg is None:
                continue
            kg = float(kg)
            if block_id in SPLIT_RATIOS:
                for sub_block_id, share in SPLIT_RATIOS[block_id].items():
                    rows.append(HistoricalAnnualYield(
                        block_id=sub_block_id, season_year=year, kg=round(kg * share, 1), estimated=True,
                    ))
            else:
                rows.append(HistoricalAnnualYield(block_id=block_id, season_year=year, kg=kg))
    return rows


def load_kg_per_blok_rows(wb):
    ws = wb[KG_PER_BLOK_SHEET]
    year_col_idx = column_index_from_string(KG_PER_BLOK_YEAR_COL) - 1
    totaal_col_idx = column_index_from_string(KG_PER_BLOK_TOTAAL_COL) - 1
    rows = []
    for row in ws.iter_rows(min_row=KG_PER_BLOK_MIN_ROW, max_row=KG_PER_BLOK_MAX_ROW, values_only=True):
        year = row[year_col_idx]
        if not isinstance(year, int) or year not in KG_PER_BLOK_YEARS:
            continue
        totaal = row[totaal_col_idx]
        if totaal is None:
            continue
        rows.append(HistoricalAnnualYield(block_id=None, season_year=year, kg=float(totaal)))
    return rows


def load_rows():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    return load_produksie_rows(wb) + load_kg_per_blok_rows(wb)


def main():
    create_db_and_tables()
    records = load_rows()
    with Session(engine) as session:
        session.exec(delete(HistoricalAnnualYield))
        session.add_all(records)
        session.commit()
    print(f"Imported {len(records)} historical annual yield rows from:\n  {XLSX_PATH}")


if __name__ == "__main__":
    main()
