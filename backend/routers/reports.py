import io
import os
from collections import Counter
from datetime import date
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlmodel import Session, select

from db import DATA_DIR, get_session
from excel_io import rows_to_xlsx_bytes
from models import Block, Device, HarvestRecord, HistoricalAnnualYield, HistoricalHarvest, Lot, LotStatus, \
    ReceivingRecord, Supplier, SystemSetting, Team, Worker
from routers.analysis import _block_sort_key
from routers.dashboard import dashboard_summary
from routers.lots import list_in_transit, list_pending, list_received
from routers.payments import _worker_ids_for_supplier
from security import get_current_admin
from timeutil import day_bounds, local_str, to_local

router = APIRouter(prefix="/api/reports", tags=["reports"])
XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

REPORTS_DIR = os.path.join(DATA_DIR, "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)


def _xlsx_response(headers, rows, sheet_title, filename):
    data = rows_to_xlsx_bytes(headers, rows, sheet_title)
    # Every generated report is also kept on disk in data/reports/, so it's
    # swept up by whatever backup routine already covers the data/ folder -
    # not just left in whichever browser downloaded it.
    with open(os.path.join(REPORTS_DIR, filename), "wb") as f:
        f.write(data)
    return Response(content=data, media_type=XLSX_MEDIA,
                     headers={"Content-Disposition": f'attachment; filename="{filename}"'})


def _mean(values: list, places: int):
    """Blank rather than 0 when nothing was measured - an empty cell reads as
    "not recorded", where 0 would read as a freezing morning."""
    if not values:
        return ""
    return round(sum(values) / len(values), places) if places else round(sum(values) / len(values))


@router.get("/daily-harvest")
def daily_harvest_report(day: date = Query(default_factory=date.today), supplier_id: Optional[int] = None,
                          session: Session = Depends(get_session), admin=Depends(get_current_admin)):
    start, end = day_bounds(day)
    query = select(HarvestRecord).where(HarvestRecord.timestamp >= start, HarvestRecord.timestamp <= end)
    worker_ids = _worker_ids_for_supplier(session, supplier_id)
    if worker_ids is not None:
        query = query.where(HarvestRecord.worker_id.in_(worker_ids))
    records = session.exec(query).all()
    blocks = {b.id: b for b in session.exec(select(Block)).all()}
    teams = {t.id: t for t in session.exec(select(Team)).all()}

    totals: dict = {}
    for r in records:
        key = (r.block_id, r.team_id)
        entry = totals.setdefault(key, {"crates": 0, "kg": 0.0, "temps": [], "humidity": [],
                                         "conditions": Counter()})
        entry["crates"] += 1
        entry["kg"] += r.weight_kg - r.deduction_kg
        # Weather is stamped per crate at check-in and is absent on crates
        # captured before that existed, or when the lookup failed - so each
        # measure is averaged over whichever crates actually carry it.
        if r.weather_temp is not None:
            entry["temps"].append(r.weather_temp)
        if r.weather_humidity is not None:
            entry["humidity"].append(r.weather_humidity)
        if r.weather_condition:
            entry["conditions"][r.weather_condition] += 1

    headers = ["Date", "Block", "Variety", "Team", "Induna", "Crates", "Kg",
               "Avg Temp (°C)", "Avg Humidity (%)", "Conditions"]
    rows = []
    for (block_id, team_id), data in sorted(totals.items(), key=lambda x: (x[0][0] or "", x[0][1] or "")):
        block = blocks.get(block_id)
        team = teams.get(team_id)
        rows.append([
            day.isoformat(), block_id or "", block.variety if block else "", team.name if team else team_id or "",
            team.induna if team else "", data["crates"], round(data["kg"], 1),
            _mean(data["temps"], 1), _mean(data["humidity"], 0),
            # Whatever it was doing for most of the picking, so a block picked
            # through a passing shower doesn't read as a clear morning.
            data["conditions"].most_common(1)[0][0] if data["conditions"] else "",
        ])
    return _xlsx_response(headers, rows, "Daily Harvest", f"Daily_Harvest_{day}.xlsx")


@router.get("/harvest-data")
def harvest_data_report(period_start: date, period_end: date, supplier_id: Optional[int] = None,
                         session: Session = Depends(get_session), admin=Depends(get_current_admin)):
    """Daaglikse Oesdata / Daily Harvest Data: the block x date pivot behind
    the paper "Daaglikse Oesdata" log - one column per block, one row per
    day, matching how the paper form and the season workbook both lay it
    out."""
    start, end = day_bounds(period_start, period_end)
    worker_ids = _worker_ids_for_supplier(session, supplier_id)
    query = select(HarvestRecord).where(HarvestRecord.timestamp >= start, HarvestRecord.timestamp <= end)
    if worker_ids is not None:
        query = query.where(HarvestRecord.worker_id.in_(worker_ids))
    records = session.exec(query).all()
    blocks = {b.id: b for b in session.exec(select(Block)).all()}

    block_days: dict = {}   # block_id -> {day: kg}
    block_totals: dict = {}  # block_id -> kg
    day_totals: dict = {}   # day -> {"kg", "workers", "crates"}
    for r in records:
        local_ts = to_local(r.timestamp)
        if local_ts is None:
            continue
        day = local_ts.date()
        kg = r.weight_kg - r.deduction_kg
        block_days.setdefault(r.block_id, {})
        block_days[r.block_id][day] = block_days[r.block_id].get(day, 0.0) + kg
        block_totals[r.block_id] = block_totals.get(r.block_id, 0.0) + kg
        dt = day_totals.setdefault(day, {"kg": 0.0, "workers": set(), "crates": 0})
        dt["kg"] += kg
        dt["crates"] += 1
        if r.worker_id:
            dt["workers"].add(r.worker_id)

    days = sorted(day_totals.keys())
    block_ids = sorted(block_totals, key=lambda bid: ((blocks.get(bid).name if blocks.get(bid) else "") or bid or ""),
                        reverse=True)
    block_names = [(blocks[bid].name if blocks.get(bid) and blocks[bid].name else (bid or "")) for bid in block_ids]

    headers = ["Date"] + block_names + \
              ["Daily Total Kg", "Number of Workers", "Avg Kg/Worker", "Number of Crates", "Avg Kg/Crate"]
    tail_blank = ["", "", "", "", ""]

    rows = [
        ["Variety"] + [blocks[bid].variety if bid in blocks else "" for bid in block_ids] + tail_blank,
        ["Trees"] + [blocks[bid].trees if bid in blocks else "" for bid in block_ids] + tail_blank,
        ["Hectares"] + [blocks[bid].hectares if bid in blocks else "" for bid in block_ids] + tail_blank,
        [""] * len(headers),
    ]
    for d in days:
        dt = day_totals[d]
        row = [d.isoformat()]
        row += [round(block_days[bid][d], 1) if d in block_days.get(bid, {}) else "" for bid in block_ids]
        row += [
            round(dt["kg"], 1), len(dt["workers"]),
            round(dt["kg"] / len(dt["workers"]), 1) if dt["workers"] else "",
            dt["crates"],
            round(dt["kg"] / dt["crates"], 1) if dt["crates"] else "",
        ]
        rows.append(row)

    grand_total_kg = round(sum(block_totals.values()), 1)
    total_trees = sum((blocks[bid].trees if bid in blocks else 0) for bid in block_ids)
    total_hectares = sum((blocks[bid].hectares if bid in blocks else 0) for bid in block_ids)

    rows.append([""] * len(headers))
    rows.append(["Total Kg"] + [round(block_totals[bid], 1) for bid in block_ids] +
                [grand_total_kg, "", "", "", ""])
    rows.append(["Avg Kg/Tree"] + [
        round(block_totals[bid] / blocks[bid].trees, 2) if bid in blocks and blocks[bid].trees else ""
        for bid in block_ids
    ] + [round(grand_total_kg / total_trees, 2) if total_trees else "", "", "", "", ""])
    rows.append(["Avg Kg/Hectare"] + [
        round(block_totals[bid] / blocks[bid].hectares, 2) if bid in blocks and blocks[bid].hectares else ""
        for bid in block_ids
    ] + [round(grand_total_kg / total_hectares, 2) if total_hectares else "", "", "", "", ""])

    return _xlsx_response(headers, rows, "Daily Harvest Data",
                           f"Daily_Harvest_Data_{period_start}_{period_end}.xlsx")


@router.get("/lot-receiving")
def lot_receiving_report(date_from: date, date_to: date, supplier_id: Optional[int] = None,
                          session: Session = Depends(get_session), admin=Depends(get_current_admin)):
    start, end = day_bounds(date_from, date_to)
    query = select(Lot).where(Lot.timestamp >= start, Lot.timestamp <= end)
    if supplier_id is not None:
        query = query.where(Lot.supplier_id == supplier_id)
    lots = session.exec(query.order_by(Lot.timestamp)).all()
    receiving_by_lot = {}
    for rec in session.exec(select(ReceivingRecord)).all():
        receiving_by_lot.setdefault(rec.lot_id, rec)
    suppliers = {s.id: s for s in session.exec(select(Supplier)).all()}

    headers = ["Slip Number", "Supplier", "Dispatched", "Team", "Driver", "Expected Crates", "Kg Sent",
               "Status", "Received At", "Actual Crates", "Discrepancy", "Condition", "Waste Kg",
               "Temp (°C)", "Humidity (%)", "Weather"]
    rows = []
    for lot in lots:
        rec = receiving_by_lot.get(lot.id)
        supplier = suppliers.get(lot.supplier_id)
        rows.append([
            lot.slip_number, supplier.name if supplier else "",
            local_str(lot.timestamp), lot.team_id, lot.driver,
            lot.total_crates, round(lot.total_kg, 1), lot.status.value,
            local_str(lot.received_at),
            rec.actual_crates if rec else "", rec.discrepancy if rec else "",
            rec.condition if rec else "", round(rec.waste_kg, 1) if rec else "",
            lot.weather_temp if lot.weather_temp is not None else "",
            lot.weather_humidity if lot.weather_humidity is not None else "",
            lot.weather_condition or "",
        ])
    return _xlsx_response(headers, rows, "Lot & Receiving", f"Lot_Receiving_{date_from}_{date_to}.xlsx")


@router.get("/picking-notes")
def picking_notes_report(date_from: date, date_to: date, supplier_id: Optional[int] = None,
                          session: Session = Depends(get_session), admin=Depends(get_current_admin)):
    start, end = day_bounds(date_from, date_to)
    query = select(Lot).where(Lot.timestamp >= start, Lot.timestamp <= end)
    if supplier_id is not None:
        query = query.where(Lot.supplier_id == supplier_id)
    lots = session.exec(query.order_by(Lot.timestamp)).all()
    receiving_by_lot = {}
    for rec in session.exec(select(ReceivingRecord)).all():
        receiving_by_lot.setdefault(rec.lot_id, rec)
    suppliers = {s.id: s for s in session.exec(select(Supplier)).all()}
    teams = {t.id: t for t in session.exec(select(Team)).all()}

    headers = ["Slip Number", "Date", "Time", "Block", "Team", "Crates Sent", "Crates Received",
               "Total Kg", "Driver", "Supplier", "Condition", "Notes", "Received By",
               "Weather", "Temp (°C)", "Humidity (%)"]
    entries = []
    for lot in lots:
        rec = receiving_by_lot.get(lot.id)
        supplier = suppliers.get(lot.supplier_id)
        team = teams.get(lot.team_id)
        # A lot has no block field of its own - it's stamped per crate, so the
        # slip's block(s) are whatever its HarvestRecords were captured against.
        crates = session.exec(select(HarvestRecord).where(HarvestRecord.lot_id == lot.id)).all()
        blocks = sorted({c.block_id for c in crates if c.block_id})
        local_ts = to_local(lot.timestamp)
        team_name = team.name if team else lot.team_id or ""
        entries.append((
            local_ts.date() if local_ts else date.min, team_name, local_ts,
            [
                lot.slip_number, local_ts.strftime("%Y-%m-%d") if local_ts else "",
                local_ts.strftime("%H:%M") if local_ts else "", ", ".join(blocks),
                team_name, lot.total_crates,
                rec.actual_crates if rec else "", round(lot.total_kg, 1), lot.driver,
                supplier.name if supplier else "",
                rec.condition if rec else "", rec.notes if rec else "", rec.received_by if rec else "",
                lot.weather_condition or "",
                lot.weather_temp if lot.weather_temp is not None else "",
                lot.weather_humidity if lot.weather_humidity is not None else "",
            ],
        ))
    entries.sort(key=lambda e: (e[0], e[1], e[2]))
    rows = [e[3] for e in entries]
    return _xlsx_response(headers, rows, "Picking Notes", f"Picking_Notes_{date_from}_{date_to}.xlsx")


@router.get("/team-picking-list")
def team_picking_list_report(date_from: date, date_to: date, supplier_id: Optional[int] = None,
                              session: Session = Depends(get_session), admin=Depends(get_current_admin)):
    """Span Pluklys / Team Picking List: one row per team per day, with the
    day's blocks (kg + deductions) and dispatched lots (crates, time, slip
    number) laid out as repeating column groups - matching the fields on the
    paper "Inligting van die Dag" slip an induna's team fills in by hand."""
    start, end = day_bounds(date_from, date_to)

    worker_ids = _worker_ids_for_supplier(session, supplier_id)
    hr_query = select(HarvestRecord).where(HarvestRecord.timestamp >= start, HarvestRecord.timestamp <= end)
    if worker_ids is not None:
        hr_query = hr_query.where(HarvestRecord.worker_id.in_(worker_ids))
    records = session.exec(hr_query).all()

    lot_query = select(Lot).where(Lot.timestamp >= start, Lot.timestamp <= end)
    if supplier_id is not None:
        lot_query = lot_query.where(Lot.supplier_id == supplier_id)
    lots = session.exec(lot_query).all()

    teams = {t.id: t for t in session.exec(select(Team)).all()}
    blocks = {b.id: b for b in session.exec(select(Block)).all()}
    devices = {d.id: d for d in session.exec(select(Device)).all()}

    groups: dict = {}

    def _group(team_id, day):
        return groups.setdefault((team_id, day), {
            "devices": Counter(), "workers": set(), "blocks": {}, "lots": [],
        })

    for r in records:
        local_ts = to_local(r.timestamp)
        day = local_ts.date() if local_ts else date.min
        g = _group(r.team_id, day)
        if r.device_id:
            g["devices"][r.device_id] += 1
        if r.worker_id:
            g["workers"].add(r.worker_id)
        entry = g["blocks"].setdefault(r.block_id, {"kg": 0.0, "deductions": 0.0})
        entry["kg"] += r.weight_kg
        entry["deductions"] += r.deduction_kg

    for lot in lots:
        local_ts = to_local(lot.timestamp)
        day = local_ts.date() if local_ts else date.min
        g = _group(lot.team_id, day)
        if not g["devices"] and lot.device_id:
            g["devices"][lot.device_id] += 1
        g["lots"].append((local_ts, lot))

    max_blocks = max((len(g["blocks"]) for g in groups.values()), default=0)
    max_lots = max((len(g["lots"]) for g in groups.values()), default=0)

    headers = ["Data Capturer", "Team", "Induna", "Date", "Workers", "Total Deductions"]
    for i in range(1, max_blocks + 1):
        headers += [f"Block {i} Name", f"Block {i} Total Kg", f"Block {i} Deductions"]
    for i in range(1, max_lots + 1):
        headers += [f"Lot {i} Crates", f"Lot {i} Time", f"Lot {i} Slip Number"]

    rows = []
    for (team_id, day), g in sorted(groups.items(), key=lambda kv: (kv[0][1], teams.get(kv[0][0]).name if teams.get(kv[0][0]) else kv[0][0] or "")):
        team = teams.get(team_id)
        device = devices.get(g["devices"].most_common(1)[0][0]) if g["devices"] else None
        total_deductions = sum(b["deductions"] for b in g["blocks"].values())
        row = [
            device.data_capturer if device else "",
            team.name if team else team_id or "", team.induna if team else "",
            day.isoformat() if day != date.min else "", len(g["workers"]), round(total_deductions, 1),
        ]
        for block_id in sorted(g["blocks"], key=lambda x: x or ""):
            block = blocks.get(block_id)
            b = g["blocks"][block_id]
            row += [block.name or block_id if block else (block_id or ""), round(b["kg"], 1), round(b["deductions"], 1)]
        row += [""] * (3 * (max_blocks - len(g["blocks"])))
        for local_ts, lot in sorted(g["lots"], key=lambda x: x[1].timestamp):
            row += [lot.total_crates, local_ts.strftime("%H:%M") if local_ts else "", lot.slip_number]
        row += [""] * (3 * (max_lots - len(g["lots"])))
        rows.append(row)

    return _xlsx_response(headers, rows, "Team Picking List", f"Team_Picking_List_{date_from}_{date_to}.xlsx")


def _lot_rows(lots_data: list) -> list:
    return [[
        l["slip_number"], l["supplier_name"], l["team_id"] or "", l["driver"], l["total_crates"], l["total_kg"],
        l["age_minutes"],
    ] for l in lots_data]


@router.get("/harvesting-list")
def harvesting_list_report(period_start: date, period_end: date, supplier_id: Optional[int] = None,
                            session: Session = Depends(get_session), admin=Depends(get_current_admin)):
    lots_data = list_pending(supplier_id=supplier_id, period_start=period_start, period_end=period_end,
                              session=session)
    headers = ["Slip Number", "Farm/Supplier", "Team", "Driver", "Crates", "Kg", "Age (min)"]
    return _xlsx_response(headers, _lot_rows(lots_data), "Harvesting",
                           f"Harvesting_{period_start}_{period_end}.xlsx")


@router.get("/in-transit-list")
def in_transit_list_report(period_start: date, period_end: date, supplier_id: Optional[int] = None,
                            session: Session = Depends(get_session), admin=Depends(get_current_admin)):
    lots_data = list_in_transit(supplier_id=supplier_id, period_start=period_start, period_end=period_end,
                                 session=session)
    headers = ["Slip Number", "Farm/Supplier", "Team", "Driver", "Crates", "Kg", "Age (min)"]
    return _xlsx_response(headers, _lot_rows(lots_data), "In Transit",
                           f"In_Transit_{period_start}_{period_end}.xlsx")


@router.get("/received-list")
def received_list_report(period_start: date, period_end: date, supplier_id: Optional[int] = None,
                          session: Session = Depends(get_session), admin=Depends(get_current_admin)):
    """Matches the packhouse's paper "Packhouse Receipt Lists" slip: date and
    time split out, plus the receiving block and rejected (waste) amount."""
    lots_data = list_received(period_start=period_start, period_end=period_end, supplier_id=supplier_id,
                               session=session)
    lot_ids = [l["id"] for l in lots_data]
    receiving_by_lot = {}
    for rec in session.exec(select(ReceivingRecord).where(ReceivingRecord.lot_id.in_(lot_ids))).all():
        receiving_by_lot.setdefault(rec.lot_id, rec)
    blocks_by_lot = {}
    for c in session.exec(select(HarvestRecord).where(HarvestRecord.lot_id.in_(lot_ids))).all():
        if c.block_id:
            blocks_by_lot.setdefault(c.lot_id, set()).add(c.block_id)

    headers = ["Slip Number", "Date", "Time", "Block", "Farm/Supplier", "Team", "Driver", "Crates", "Kg",
               "Rejected"]
    rows = []
    for l in lots_data:
        rec = receiving_by_lot.get(l["id"])
        local_ts = to_local(l["received_at"])
        rows.append([
            l["slip_number"],
            local_ts.strftime("%Y-%m-%d") if local_ts else "",
            local_ts.strftime("%H:%M") if local_ts else "",
            ", ".join(sorted(blocks_by_lot.get(l["id"], []))),
            l["supplier_name"], l["team_id"] or "", l["driver"], l["total_crates"], l["total_kg"],
            round(rec.waste_kg, 1) if rec else "",
        ])
    return _xlsx_response(headers, rows, "Received", f"Received_{period_start}_{period_end}.xlsx")


@router.get("/worker-harvest")
def worker_harvest_report(period_start: date, period_end: date, supplier_id: Optional[int] = None,
                           session: Session = Depends(get_session), admin=Depends(get_current_admin)):
    summary = dashboard_summary(period_start, period_end, supplier_id, session, admin)
    headers = ["Emp Nr", "Name", "Farm/Supplier", "Crates", "Kg", "Amount Due", "Avg Kg/Crate"]
    rows = [[
        w["worker_id"], w["name"], w["supplier_name"], w["crates"], w["total_kg"], w["amount_due"],
        w["avg_kg_crate"],
    ] for w in summary["workers"]]
    return _xlsx_response(headers, rows, "Worker Harvest", f"Worker_Harvest_{period_start}_{period_end}.xlsx")


@router.get("/litchi-wages")
def litchi_wages_report(period_start: date, period_end: date, supplier_id: Optional[int] = None,
                         session: Session = Depends(get_session), admin=Depends(get_current_admin)):
    """Lietsjie Lone / Litchi Wages: one row per worker, crates harvested vs.
    crates actually received at the pack house broken out per day - the gap
    flags fruit that never made it off the lot it was picked into, before
    wages get paid out on it."""
    start, end = day_bounds(period_start, period_end)
    worker_ids = _worker_ids_for_supplier(session, supplier_id)
    query = select(HarvestRecord).where(HarvestRecord.timestamp >= start, HarvestRecord.timestamp <= end)
    if worker_ids is not None:
        query = query.where(HarvestRecord.worker_id.in_(worker_ids))
    records = session.exec(query).all()

    lot_ids = {r.lot_id for r in records if r.lot_id}
    lots = {l.id: l for l in session.exec(select(Lot).where(Lot.id.in_(lot_ids))).all()} if lot_ids else {}
    workers = {w.id: w for w in session.exec(select(Worker)).all()}

    worker_days: dict = {}  # worker_id -> day -> {harvested, received, deductions}
    for r in records:
        if not r.worker_id:
            continue
        local_ts = to_local(r.timestamp)
        day = local_ts.date() if local_ts else date.min
        entry = worker_days.setdefault(r.worker_id, {}).setdefault(
            day, {"harvested": 0, "received": 0, "deductions": 0.0})
        entry["harvested"] += 1
        entry["deductions"] += r.deduction_kg
        lot = lots.get(r.lot_id) if r.lot_id else None
        if lot and lot.status == LotStatus.received:
            entry["received"] += 1

    days = sorted({d for wd in worker_days.values() for d in wd})
    headers = ["Emp Nr", "Name & Surname", "ID Number"]
    for d in days:
        headers += [f"{d.isoformat()} Crates Harvested", f"{d.isoformat()} Crates Received",
                    f"{d.isoformat()} Deductions (kg)", f"{d.isoformat()} Difference"]
    headers += ["Total Crates Harvested", "Total Crates Received", "Total Deductions (kg)", "Total Difference",
                "Bank", "Account"]

    rows = []
    for worker_id in sorted(worker_days.keys()):
        w = workers.get(worker_id)
        wd = worker_days[worker_id]
        row = [worker_id, w.name if w else "", w.id_number if w else ""]
        tot_harvested = tot_received = 0
        tot_deductions = 0.0
        for d in days:
            entry = wd.get(d, {"harvested": 0, "received": 0, "deductions": 0.0})
            row += [entry["harvested"], entry["received"], round(entry["deductions"], 1),
                    entry["harvested"] - entry["received"]]
            tot_harvested += entry["harvested"]
            tot_received += entry["received"]
            tot_deductions += entry["deductions"]
        row += [tot_harvested, tot_received, round(tot_deductions, 1), tot_harvested - tot_received,
                 w.bank if w else "", w.account if w else ""]
        rows.append(row)

    return _xlsx_response(headers, rows, "Litchi Wages", f"Litchi_Wages_{period_start}_{period_end}.xlsx")


@router.get("/block-harvest")
def block_harvest_report(period_start: date, period_end: date, supplier_id: Optional[int] = None,
                          session: Session = Depends(get_session), admin=Depends(get_current_admin)):
    summary = dashboard_summary(period_start, period_end, supplier_id, session, admin)
    headers = ["Block", "Crates", "Kg", "Avg Kg/Crate", "Avg Kg/Tree", "Avg Kg/Ha"]
    rows = [[
        b["name"], b["crates"], b["total_kg"], b["avg_kg_crate"],
        b["avg_kg_tree"] if b["avg_kg_tree"] is not None else "",
        b["avg_kg_hectare"] if b["avg_kg_hectare"] is not None else "",
    ] for b in summary["blocks"]]
    return _xlsx_response(headers, rows, "Block Harvest", f"Block_Harvest_{period_start}_{period_end}.xlsx")


def _style_header_cell(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2E7D32")
    cell.alignment = Alignment(horizontal="center", vertical="center")


@router.get("/historical-harvest-data")
def historical_harvest_data_report(session: Session = Depends(get_session), admin=Depends(get_current_admin)):
    """Historical Harvest Data: every harvest figure this farm has on file,
    1987 through the current season, in one workbook. Not date-range
    filtered like the other reports - there's only ever one of these.

    Three different grains of record sit side by side here, which is why
    there are several sheets rather than one big grid:
      - Per-year sheets (2020 on): the full block x date pivot - the farm's
        own "Daaglikse Oesdata" workbook, kept live. Pre-app seasons come
        from HistoricalHarvest (imported once - see
        scripts/import_historical_harvest.py for provenance and the
        block-split-by-hectare-ratio caveat); the current season is built
        fresh from HarvestRecord on every download, so it's always up to
        date without a re-import.
      - Annual Totals (1987-2019): season totals only, no daily breakdown -
        per block back to 2012, whole-farm-only before that (see
        scripts/import_historical_annual_yield.py).
      - Season Summary and Block by Year: cross-era views built from
        whichever of the above covers each season, so the whole record can
        be read at once. Season Summary names each season's grain in its
        own column so the two are never silently mixed."""
    settings = session.exec(select(SystemSetting)).first()
    current_year = settings.current_harvest_year if settings else date.today().year
    blocks = {b.id: b for b in session.exec(select(Block)).all()}

    day_kg: dict = {}  # (year, block_id, date) -> kg
    estimated_blocks: set = set()
    for h in session.exec(select(HistoricalHarvest)).all():
        key = (h.season_year, h.block_id, h.harvest_date)
        day_kg[key] = day_kg.get(key, 0.0) + h.kg
        if h.estimated:
            estimated_blocks.add(h.block_id)
    for r in session.exec(select(HarvestRecord)).all():
        local_ts = to_local(r.timestamp)
        if local_ts is None or local_ts.year != current_year:
            continue
        key = (current_year, r.block_id, local_ts.date())
        day_kg[key] = day_kg.get(key, 0.0) + (r.weight_kg - r.deduction_kg)

    years = sorted({k[0] for k in day_kg})

    annual_kg: dict = {}  # (year, block_id) -> kg
    annual_estimated_blocks: set = set()
    for a in session.exec(select(HistoricalAnnualYield)).all():
        key = (a.season_year, a.block_id)
        annual_kg[key] = annual_kg.get(key, 0.0) + a.kg
        if a.estimated:
            annual_estimated_blocks.add(a.block_id)
    annual_years = sorted({k[0] for k in annual_kg})

    def block_label(bid, estimated_set=estimated_blocks):
        b = blocks.get(bid)
        name = b.name if b else (bid or "")
        return f"{name}*" if bid in estimated_set else name

    # One per-block annual figure per (year, block), however that year was
    # recorded: summed from the daily sheets where those exist (2020 on),
    # taken straight from the annual-only import before that. The 1987-2009
    # rows carry block_id None - a whole-farm total with no block breakdown
    # (see HistoricalAnnualYield) - so they're deliberately excluded here
    # and appear only in the Season Summary's total.
    block_year_kg: dict = {}  # (year, block_id) -> kg
    for (year, block_id, _), kg in day_kg.items():
        block_year_kg[(year, block_id)] = block_year_kg.get((year, block_id), 0.0) + kg
    for (year, block_id), kg in annual_kg.items():
        if block_id is not None:
            block_year_kg[(year, block_id)] = block_year_kg.get((year, block_id), 0.0) + kg

    all_years = sorted({y for y, _ in block_year_kg} | set(annual_years) | set(years))
    all_estimated = estimated_blocks | annual_estimated_blocks

    # How each season was recorded, so a reader can tell a real daily record
    # from a single hand-written season total - they are not equally solid.
    def granularity(year):
        if year in years:
            return "Daily, per block" if year != current_year else "Daily, per block (in progress)"
        if any((year, bid) in annual_kg for bid in blocks):
            return "Season total, per block"
        return "Season total, whole farm only"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Blocks reference sheet ---
    bs = wb.create_sheet("Blocks")
    bs.append(["Id", "Name", "Variety", "Trees", "Hectares", "Active"])
    for c in bs[1]:
        _style_header_cell(c)
    for block_id in sorted(blocks, key=_block_sort_key):
        b = blocks[block_id]
        bs.append([b.id, b.name, b.variety, b.trees, b.hectares, "Active" if b.active else "Inactive"])
    bs.freeze_panes = "A2"
    bs.column_dimensions["B"].width = 16

    # --- Notes sheet ---
    ns = wb.create_sheet("Notes")
    ns.column_dimensions["A"].width = 100
    notes = [
        "Notes on this report",
        "",
        f"Generated {date.today().isoformat()}. Historical seasons ({years[0] if years else '-'}-"
        f"{current_year - 1}) are a fixed, one-time import from the farm's pre-app records. The current "
        f"season ({current_year}) sheet is generated fresh from live harvest data every time this report "
        "is downloaded.",
        "",
        "Block splits: a handful of today's blocks (8a/8b, 10a/10b, 17a/17b, 19a/19b) didn't exist "
        "separately before this app - the original records had one combined daily total for each pair. "
        "Those historical totals were split between the two sub-blocks in proportion to their hectares. "
        "This is an ESTIMATE, not an actually recorded per-sub-block figure - affected block names are "
        "marked with an asterisk (*) in each year's sheet.",
        "",
        "Figures are in kg per block per day.",
        "",
        "Season Summary lists every season on file with its total and how it was recorded - a real "
        "day-by-day record and a single hand-written season total are both here, and the "
        "\"How It Was Recorded\" column says which is which. Block by Year puts every block-level "
        "season total in one grid, whichever way that year was recorded, so blocks can be compared "
        "across the years side by side; seasons with no block breakdown at all (1987-2009) are "
        "counted in Season Summary but can't appear there.",
    ]
    if annual_years:
        notes += [
            "",
            f"Annual Totals sheet ({annual_years[0]}-{annual_years[-1]}): the farm's older records only "
            "kept totals per SEASON, not per day, so these years have no daily breakdown and aren't part "
            "of the Analysis tab or Risk indicator (which need day-by-day figures, and weather data only "
            "goes back to 2020 anyway). The same block-split estimate and asterisk convention above "
            "applies to the per-block years here too.",
            "",
            "Rows marked \"whole-farm total only\" in that sheet's Notes column go back further still "
            "(1987-2009) to records that predate today's block register entirely, under a completely "
            "different, incompatible block-numbering scheme. Rather than guess at a mapping between old "
            "and new block numbers, only each year's whole-farm total was kept for those rows - no "
            "per-block breakdown is available.",
        ]
    for i, line in enumerate(notes, start=1):
        cell = ns.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # --- Season Summary: one row per season, every year on file ---
    if all_years:
        ss = wb.create_sheet("Season Summary")
        ss.append(["Year", "Total Kg", "Change vs Previous", "Blocks Recorded", "How It Was Recorded"])
        for c in ss[1]:
            _style_header_cell(c)
        season_total = {}
        for year in all_years:
            if year in years:
                season_total[year] = sum(kg for (y, _, _), kg in day_kg.items() if y == year)
            elif (year, None) in annual_kg:
                season_total[year] = annual_kg[(year, None)]
            else:
                season_total[year] = sum(kg for (y, bid), kg in annual_kg.items()
                                          if y == year and bid is not None)
        for year in all_years:
            total = season_total[year]
            block_count = len({bid for (y, bid) in block_year_kg if y == year and bid is not None})
            # Only against the season immediately before - the record has
            # gaps (nothing for 2010-2011), and calling a three-year jump a
            # year-on-year change would read as a collapse or a boom that
            # never happened.
            prev = season_total.get(year - 1)
            change = (total - prev) / prev if prev else None
            ss.append([year, round(total, 1), change, block_count or None, granularity(year)])
        for r in range(2, ss.max_row + 1):
            ss.cell(row=r, column=3).number_format = "+0.0%;-0.0%"
        ss.freeze_panes = "A2"
        for col, width in zip("ABCDE", (8, 14, 18, 16, 30)):
            ss.column_dimensions[col].width = width

    # --- Block by Year: every block-level season total in one grid ---
    if block_year_kg:
        by_years = sorted({y for y, _ in block_year_kg})
        by_blocks = sorted({bid for _, bid in block_year_kg if bid is not None}, key=_block_sort_key)
        bys = wb.create_sheet("Block by Year")
        bys.append(["Block", "Variety", "Trees", "Hectares"] + [str(y) for y in by_years] + ["Total"])
        for c in bys[1]:
            _style_header_cell(c)
        for bid in by_blocks:
            b = blocks.get(bid)
            row = [block_label(bid, all_estimated), b.variety if b else "",
                   b.trees if b else None, b.hectares if b else None]
            vals = [block_year_kg.get((y, bid)) for y in by_years]
            row += [round(v, 1) if v is not None else None for v in vals]
            row.append(round(sum(v for v in vals if v is not None), 1))
            bys.append(row)
        totals_row = ["TOTAL", "", None, None]
        for y in by_years:
            totals_row.append(round(sum(kg for (yy, bid), kg in block_year_kg.items()
                                         if yy == y and bid is not None), 1))
        totals_row.append(round(sum(kg for (_, bid), kg in block_year_kg.items() if bid is not None), 1))
        bys.append(totals_row)
        for c in bys[bys.max_row]:
            c.font = Font(bold=True)
        bys.freeze_panes = "E2"
        bys.column_dimensions["A"].width = 16
        bys.column_dimensions["B"].width = 12
        for col in range(3, len(by_years) + 6):
            bys.column_dimensions[get_column_letter(col)].width = 10

    # --- Annual Totals sheet (even older, season-only figures) ---
    if annual_years:
        # block_id None marks a whole-farm-only year (no block breakdown available that far back)
        annual_blocks = sorted({bid for (_, bid) in annual_kg if bid is not None}, key=_block_sort_key)
        farm_total_years = {y for (y, bid) in annual_kg if bid is None}
        as_ = wb.create_sheet("Annual Totals")
        header = (["Year"] + [block_label(bid, annual_estimated_blocks) for bid in annual_blocks] +
                   ["Total", "Notes"])
        as_.append(header)
        for c in as_[1]:
            _style_header_cell(c)
        for year in annual_years:
            if year in farm_total_years:
                total = annual_kg[(year, None)]
                as_.append([year] + [None] * len(annual_blocks) + [round(total, 1), "whole-farm total only"])
            else:
                row_kg = {bid: annual_kg.get((year, bid), 0.0) for bid in annual_blocks}
                total = sum(row_kg.values())
                as_.append([year] + [round(row_kg[bid], 1) if (year, bid) in annual_kg else None
                                      for bid in annual_blocks] + [round(total, 1), ""])
        as_.freeze_panes = "B2"
        as_.column_dimensions["A"].width = 8
        for col in range(2, len(header)):
            as_.column_dimensions[get_column_letter(col)].width = 11
        as_.column_dimensions[get_column_letter(len(header))].width = 20

    # --- Per-year sheets ---
    for year in years:
        year_days: dict = {}  # date -> {block_id: kg}
        year_blocks: set = set()
        for (y, block_id, d), kg in day_kg.items():
            if y != year:
                continue
            bucket = year_days.setdefault(d, {})
            bucket[block_id] = bucket.get(block_id, 0.0) + kg
            year_blocks.add(block_id)
        block_ids = sorted(year_blocks, key=_block_sort_key)

        ws = wb.create_sheet(str(year))
        header = ["Date", "Weekday"] + [block_label(bid) for bid in block_ids] + ["Total"]
        ws.append(header)
        for c in ws[1]:
            _style_header_cell(c)

        for d in sorted(year_days):
            row_kg = year_days[d]
            total = sum(row_kg.get(bid, 0.0) for bid in block_ids)
            ws.append([d, d.strftime("%A")] + [round(row_kg.get(bid, 0.0), 1) for bid in block_ids] +
                      [round(total, 1)])

        first_data_row = 2
        last_data_row = ws.max_row
        ws.append(["", "TOTAL"])
        footer_row = ws.max_row
        for i in range(len(block_ids) + 1):
            col_idx = 3 + i
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=footer_row, column=col_idx,
                            value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})")
            cell.font = Font(bold=True)
        ws.cell(row=footer_row, column=2).font = Font(bold=True)

        for r in range(2, last_data_row + 1):
            ws.cell(row=r, column=1).number_format = "dd/mm/yyyy"

        ws.freeze_panes = "C2"
        for col in range(1, len(header) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 11
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 11

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    # Span the whole workbook, not just the daily sheets - it reaches back
    # to the earliest season-total year, well before the first daily one.
    filename = (f"Historical_Harvest_Data_{all_years[0]}_{all_years[-1]}.xlsx"
                if all_years else "Historical_Harvest_Data.xlsx")
    with open(os.path.join(REPORTS_DIR, filename), "wb") as f:
        f.write(data)
    return Response(content=data, media_type=XLSX_MEDIA,
                     headers={"Content-Disposition": f'attachment; filename="{filename}"'})
