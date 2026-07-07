"""Generic CSV/.xlsx helpers used by master-data import/export and reports.

Keeping this generic (headers + rows in, bytes out / dicts out) means every
router just supplies plain Python data - no per-report boilerplate for the
actual file format.
"""
import csv
import io
from typing import Any

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def rows_to_xlsx_bytes(headers: list[str], rows: list[list[Any]], sheet_title: str = "Sheet1") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel sheet name limit
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for i, header in enumerate(headers, start=1):
        width = max(len(str(header)), *(len(str(r[i - 1])) for r in rows)) if rows else len(str(header))
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 40)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def rows_to_csv_bytes(headers: list[str], rows: list[list[Any]]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")  # BOM so Excel opens UTF-8 correctly


async def parse_uploaded_table(file: UploadFile) -> list[dict]:
    """Returns a list of dict rows keyed by the header row. Accepts .csv or .xlsx."""
    content = await file.read()
    name = (file.filename or "").lower()
    if name.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        return [dict(zip(headers, row)) for row in rows_iter if any(v is not None for v in row)]
    else:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [row for row in reader]
